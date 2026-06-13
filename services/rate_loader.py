from collections.abc import Sequence
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from domain.rate.model import RateCell, RateTableVersion
from integrations.db.repositories import product_repo, rate_repo
from observability.exceptions import ConflictError, NotFoundError, ValidationError
from services.rating import canonical_dim_key

_RATE_COLUMN = "gross premium before discount"
_AUDIT_COLUMNS = {
    "gross_after_discount": "gross premium after discount",
    "net_premium": "net premium",
    "discount": "discount",
}


def _find_header_row(rows: Sequence[Sequence], declared_names: Sequence[str]) -> int:
    """Locate the header row by matching all declared dimension labels
    (case-insensitive). Tolerates leading blank rows in the sheet."""
    wanted = {n.lower() for n in declared_names}
    for idx, row in enumerate(rows):
        labels = {str(c).strip().lower() for c in row if c is not None}
        if wanted.issubset(labels):
            return idx
    raise ValidationError(
        f"Header row with dimensions {sorted(declared_names)} not found in sheet"
    )


def _dec(row: Sequence, col_index: dict[str, int], col_name: str) -> Decimal | None:
    i = col_index.get(col_name)
    if i is None or i >= len(row) or row[i] is None:
        return None
    return Decimal(str(row[i]))


async def load_rate_table(
    session: AsyncSession,
    *,
    product_code: str,
    xlsx_path: str,
    source_ref: str,
    effective_from: date,
    effective_to: date | None = None,
    unit_basis: Decimal = Decimal("1000"),
) -> RateTableVersion:
    """Load an xlsx rate table as a new immutable version for a product.

    Validates the sheet columns against the product's DECLARED dimensions,
    converts every value through ``Decimal(str(...))`` (never via float), refuses
    a window that overlaps an existing active version, and writes the version +
    all cells in one transaction.
    """
    product = await product_repo.get_by_code(session, product_code)
    if product is None:
        raise NotFoundError(f"Unknown product: {product_code}")
    if not product.dimensions:
        raise ValidationError(f"Product {product_code} declares no rating dimensions")

    dims = list(product.dimensions)  # ordered by position
    dim_names = [d.name for d in dims]
    dim_types = {d.name: d.data_type for d in dims}

    path = Path(xlsx_path)
    if not path.exists():
        raise ValidationError(f"Rate file not found: {xlsx_path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
    finally:
        wb.close()

    header_idx = _find_header_row(rows, dim_names)
    header = [str(c).strip().lower() if c is not None else "" for c in rows[header_idx]]
    col_index = {label: i for i, label in enumerate(header) if label}

    missing = [n for n in dim_names if n.lower() not in col_index]
    if missing:
        raise ValidationError(f"xlsx is missing dimension columns: {missing}")
    if _RATE_COLUMN not in col_index:
        raise ValidationError(f"xlsx is missing the '{_RATE_COLUMN}' column")

    clashes = await rate_repo.overlapping_active_versions(
        session, product.id, effective_from, effective_to
    )
    if clashes:
        raise ConflictError(
            f"An active rate version already overlaps "
            f"[{effective_from}, {effective_to}] for {product_code}"
        )

    version = await rate_repo.save_version(
        session,
        RateTableVersion(
            product_id=product.id,
            source_ref=source_ref,
            unit_basis=unit_basis,
            effective_from=effective_from,
            effective_to=effective_to,
            status="active",
            cell_count=0,
        ),
    )

    first_dim_col = col_index[dim_names[0].lower()]
    cells: list[RateCell] = []
    seen_keys: set[str] = set()
    for row in rows[header_idx + 1 :]:
        if row is None or all(c is None for c in row):
            continue
        if first_dim_col >= len(row) or row[first_dim_col] is None:
            continue

        dimensions: dict = {}
        for name in dim_names:
            raw = row[col_index[name.lower()]]
            dimensions[name] = (
                int(raw) if dim_types[name] == "int" else str(raw).strip().upper()
            )

        dim_key = canonical_dim_key(dimensions, dims)
        if dim_key in seen_keys:
            raise ValidationError(f"Duplicate rate cell for {dimensions}")
        seen_keys.add(dim_key)

        rate = _dec(row, col_index, _RATE_COLUMN)
        if rate is None:
            raise ValidationError(f"Missing '{_RATE_COLUMN}' value for {dimensions}")

        cells.append(
            RateCell(
                rate_table_version_id=version.id,
                dimensions=dimensions,
                dim_key=dim_key,
                gross_before_discount=rate,
                gross_after_discount=_dec(row, col_index, _AUDIT_COLUMNS["gross_after_discount"]),
                net_premium=_dec(row, col_index, _AUDIT_COLUMNS["net_premium"]),
                discount=_dec(row, col_index, _AUDIT_COLUMNS["discount"]),
            )
        )

    if not cells:
        raise ValidationError("No data rows found in rate file")

    await rate_repo.bulk_add_cells(session, cells)
    version.cell_count = len(cells)
    await session.flush()
    return version
